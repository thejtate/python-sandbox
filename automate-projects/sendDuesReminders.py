#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status. 
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each members payment status
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
	payment = sheet.cell(row=r, column=lastCol).value
	if payment != 'paid':
		name = sheet.cell(row=r, column=1).value
		email = sheet.cell(row=r, column=2).value
		unpaidMembers[name] = email

# Log in to email account. 
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('thedarkverge@gmail.com',sys.argv[1])

# Send out reminder emails. 
for name, email in unpaidMembers.items():
	body = "Subject: %s dues unpaid. \nDear %s, \nRecords show that you need to pay your damn bill for the month %s."%(latestMonth, name, latestMonth)
	print('Sending mail to %s...'%email)
	sendmailStatus = smtpObj.sendmail('thedarkverge@gmail.com', email, body)
	if sendmailStatus != {}:
		print('There was a problem sending email to %s: %s'%(email, sendmailStatus))

smtpObj.quit()